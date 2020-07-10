# - * - coding: utf-8 - * -
import os

from openpyxl import load_workbook, Workbook
import openpyxl
import time
from textfile_solution import PROJECT_PATH

def load_data_test():
    from openpyxl import load_workbook
    wb = load_workbook("./file_location/template.xlsx")  # 打开一个 xlsx 文件
    print(wb.sheetnames)
    sheet = wb.get_sheet_by_name("Sheet3")  # 看看打开的 Excel 表里面有哪些 Sheet
    # 下面读取指定的 Sheet
    print(sheet["C"])
    print(sheet["4"])
    print(sheet["C4"].value)  # c4     &lt;- C4 格的值
    print(sheet.max_row)  # 10     &lt;-最大行数
    print(sheet.max_column)  # 5     &lt;-最大列数
    for i in sheet["C"]:
        print(i.value, end=" ")  # c1 c2 c3 c4 c5 c6 c7 c8 c9 c10     &lt;-C 列中的所有值


def export_data_test():
    ls = [
        ["马坡", "接入交换", "192.168.1.1", "G0 / 3", "AAAA - AAAA - AAAA"],
        ["马坡", "接入交换", "192.168.1.2", "G0 / 8","BBBB - BBBB - BBBB"],
        ["马坡", "接入交换", "192.168.1.2", "G0 / 8", "CCCC - CCCC - CCCC"],
        ["马坡", "接入交换", "192.168.1.2", "G0 / 8", "DDDD - DDDD - DDDD"]
    ]

    # 定义数据 

    time_format = "%Y-%m-%d__%H:%M:%S"
    time_current = time.strftime(time_format)
    filename = PROJECT_PATH+"//file_location//export_template.xlsx"
    sheetname = "Sheet1"
    # 定义时间格式
    if not os.path.exists(filename):
        wb = Workbook()
        _ = wb.create_sheet(sheetname)
        wb.save(filename)
    savetoexcel(ls,sheetname, filename, time_current)


def savetoexcel(data, sheetname, wbname, time_current):
    print("写入excel：")
    wb = openpyxl.load_workbook(filename = wbname)
    # 打开 Excel 文件 

    sheet = wb.active  # 关联 Excel 活动的 Sheet(这里关联的是 Sheet1)
    max_row = sheet.max_row  # 获取 Sheet1 中最大的行数 
    row = max_row + 3  # 将新数据写入最大行数加 3 的位置 
    data_len = row + len(data)  # 计算当前数据的长度 

    for data_row in range(row,data_len):  # 写入数据 
        # 轮询每一行进行写入数据 
        for data_col1 in range(2, 7):
            # 针对每一行,下面还要执行 for 循环来写入列的数据 
            _ = sheet.cell(row=data_row, column = 1, value = str(time_current))
            # 在每行的第一列写入时间 
            _ = sheet.cell(row=data_row,column = data_col1,value = str(data[data_row - data_len]
            [data_col1 - 2]))
            # 从第二列开始写入数据
            wb.save(filename = wbname)  # 保存数据 
            print("保存成功")


def search_data_test():
    import openpyxl
    wb = openpyxl.load_workbook("./file_location/export_template.xlsx")
    the_list = []
    while True:
        info = input("请输入关键字：").upper().strip()
        if len(info) == 0:  # 输入的关键字不能为空，否则继续循环
            continue
        count = 0
        for line1 in wb["Sheet"].values:  # 轮询列表
            if None not in line1:
            ##Excel 中空行的数据表示 None，当这里匹配 None 时就不会再进行 for 循环，所以需要匹配非
            # None的数据才能进行下面的for 循环
                for line2 in line1:  # 因为列表中还存在元组，所以需要将元组的内容也轮询一遍
                    if info in line2:
                        count += 1  # 统计关键字被匹配了多少次
                        print(line1)  # 匹配关键字后输出元组信息
                    else:
                        print("匹配%s的数量统计：%s个条目被匹配"%(info, count))  # 显示查找的关键字被匹配
                        # 了多少次


def create_chart_test():
    from openpyxl import Workbook
    from openpyxl.chart import (
        AreaChart,
        Reference,
        Series,
    )

    wb = Workbook()
    ws = wb.active

    rows = [
        ['Number', 'Batch 1', 'Batch 2'],
        [2, 40, 30],
        [3, 40, 25],
        [4, 50, 30],
        [5, 30, 10],
        [6, 25, 5],
        [7, 50, 10],
    ]

    for row in rows:
        ws.append(row)

    chart = AreaChart()
    chart.title = "Area Chart"
    chart.style = 13
    chart.x_axis.title = 'Test'
    chart.y_axis.title = 'Percentage'

    cats = Reference(ws, min_col=1, min_row=1, max_row=7)
    data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=7)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    ws.add_chart(chart, "A10")

    wb.save("./file_location/area.xlsx")

if __name__ == "__main__":
    load_data_test()
    export_data_test()
    search_data_test()
    create_chart_test()